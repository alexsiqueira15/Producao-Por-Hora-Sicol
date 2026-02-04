from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
import uuid
import os
from openpyxl import Workbook
from unidecode import unidecode
from io import BytesIO
import time

# === Flask configurado para Vercel ==========================
app = Flask(
    __name__,
    template_folder="../templates",
    static_folder="../static"
)

# === Pasta temporária (filesystem efêmero da Vercel) =======
TEMP_FOLDER = "/tmp/temp_files"
os.makedirs(TEMP_FOLDER, exist_ok=True)

# === Função de limpeza automática ==========================
def limpar_temp_files():
    agora = time.time()
    for f in os.listdir(TEMP_FOLDER):
        caminho = os.path.join(TEMP_FOLDER, f)
        if os.path.isfile(caminho) and agora - os.path.getmtime(caminho) > 3600:
            os.remove(caminho)

# === Normalização de nomes =================================
def normalize_name(name):
    if not isinstance(name, str):
        return ""
    return unidecode(name.strip().upper())

# === Leitura universal de arquivos =========================
def read_any_file(file):
    content = file.read()
    file.seek(0)
    filename = file.filename.lower()

    try:
        if filename.endswith(".xlsx"):
            return pd.read_excel(BytesIO(content), engine="openpyxl")
        if filename.endswith(".xls"):
            return pd.read_excel(BytesIO(content), engine="xlrd")
    except Exception:
        pass

    try:
        return pd.read_excel(BytesIO(content), engine="openpyxl")
    except Exception:
        pass

    try:
        tables = pd.read_html(BytesIO(content))
        if tables:
            return tables[0]
    except Exception:
        pass

    try:
        return pd.read_csv(BytesIO(content), sep=None, engine="python")
    except Exception:
        pass

    raise ValueError("Não foi possível ler o arquivo.")

# === Operadores ============================================
variavel_operadores = {
    "MANHÃ": [
        "ACLECIO FERNADO MELO", "ALECIA CRISTINA EVYNA SANTOS DE JESUS", "ALISSON BRENO CUNHA DE BARROS",
        "ALLANYA GABRIELLA DE ALMEIDA SOUSA", "ANA ACACIA ARAGAO BONFIM", "BEATRIZ VITORIA NASCIMENTO DA SILVA",
        "DIOGO SANTOS MACHADO", "DEBORA MESSIAS SANTOS", "ELAINE PINTO DE CARVALHO FREIRE",
        "JONAS LUCAS DOS SANTOS", "KELLY CRISTINA TELES DOS SANTOS",
        "LUCAS MICHAEL SANTA RITA DOS SANTOS", "LUIZ MIGUEL PATRICIO FELIX", "MARCLYS ANGELICA FERREIRA SANTOS",
        "MARICLEIDE DOS SANTOS DE SOUZA", "MONIQUE ALVES DA SILVA SANTOS",
        "ROSE KATIUSKA DOS SANTOS BIGI", "ROZEANE OLIVEIRA DOS SANTOS",
        "NATALIA FARIAS SANTOS", "ADRIANA BATISTA SILVA", "MICHELE DOS SANTOS TAVARES",
        "ISRAEL SANTOS DA SILVA", "LAIS DANIELLE HONORATO ALVES SENA",
        "TAYS MILENA BISPO DOS SANTOS", "MARIA FLORENTINA MELO SANTOS",
        "LUSINEIDE SANTIAGO SANTOS", "LARISSA VITOR DA SILVA", "BEATRIZ DE ANDRADE SANTANA",
        "KEMELLY KAROLAINNY DOS SANTOS SILVA", "LUCIANE DOS SANTOS SILVA",
        "NAIARA ANDRADE SANTOS", "SILMARA SANTOS DE JESUS",
        "CRISLAINE BORGES DOS SANTOS NASCIMENTO", "HERSLANDER JORGE DOS SANTOS",
        "LORENA FERREIRA SANTOS", "ALESSANDRA SANTANA SANTOS", "JANISSON SANTANA SANTOS"
    ],
    "TARDE": [
        "ANA RAQUEL RIBEIRO DE OLIVEIRA SANTOS", "DAYANE NUNES VICENTE", "JULIA CONCEICAO SILVA SANTOS",
        "KETLYN JULIANE SILVA BATISTA", "TALITA PRISCILA FARIAS SANTOS", "DANIELLE FERNANDES BARROS SANTOS",
        "VANESSA DE OLIVEIRA GONZAGA",
        "ACUCENA DA SILVA MARCOLINO ANDRADE", "ALISSON DA CRUZ FERREIRA",
        "GABRIELLE SALOMAO DOS SANTOS", "ROBERTH SANTOS CONCEICAO",
        "TACIANE GRACIELE INACIO DOS SANTOS", "LARISSA LUIZA VIEIRA MOTA", "ACACIA DOS SANTOS CONCEICAO", 
        "ANNA BEATRIZ OLIVEIRA FONTES",
        "JAILA RODRIGUES DOS SANTOS", "EVELYN DE FREITAS SANTOS",
        "JENNIFER CAMPOS BENIGNO", "JULIANE BARROS DA SILVA",
        "LORENA MARTINS DE SANTANA", "ANA VIRGINIA SANTOS DANTAS",
        "BRUNA CAROLINA DOS SANTOS MATOS", "PAULA ROBERTA DE JESUS SANTOS",
        "VYVYAN ADRYENE SANTOS DA SILVA"
    ],
    "INTERMEDIÁRIO": [
        "GLILMA MARIA FARIAS DE MENEZES ISMERIM", "LARISSA XAVIER DE MELO",
        "RAELLY SANTOS CALDAS LIMA", "RAYARA SANTANA LIMA", "RANYSSA RAYARA DOS SANTOS", "YASMIN SANTOS DA CONCEICAO"
    ],
    "RECEPTIVO": [
        "LEONICE PAIXAO BATISTA", "EDSON XAVIER DE BRITO", "EDUARDO MOURA SANTOS", "RAINELDES GUILHERME DOS SANTOS"
    ]
}

# === ROTAS ==================================================

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/analisar", methods=["POST"])
def analisar():
    limpar_temp_files()

    if "planilha" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado."})

    file = request.files["planilha"]
    equipe = request.form.get("equipe")

    try:
        df = read_any_file(file)
    except Exception as e:
        return jsonify({"error": str(e)})

    if equipe != "variavel":
        return jsonify({"error": "Equipe inválida."})

    nome_col = "NOME SOLICITANTE"
    cpf_col = "CPF"

    if nome_col not in df.columns or cpf_col not in df.columns:
        return jsonify({"error": "Colunas obrigatórias não encontradas."})

    df[nome_col] = df[nome_col].apply(normalize_name)

    operadores_por_turno = {
        t: [normalize_name(o) for o in ops]
        for t, ops in variavel_operadores.items()
    }

    resultados = {}
    total_geral = 0

    for turno, operadores in operadores_por_turno.items():
        resultados[turno] = {}
        for operador in operadores:
            qtd = df[df[nome_col] == operador][cpf_col].nunique()
            if qtd > 0:
                resultados[turno][operador] = qtd
                total_geral += qtd

    # === Gera Excel =========================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Acordos"

    row = 1
    for turno, operadores in resultados.items():
        ws.cell(row=row, column=1, value=turno)
        row += 1
        total_turno = 0
        for operador, qtd in operadores.items():
            ws.cell(row=row, column=1, value=operador)
            ws.cell(row=row, column=2, value=qtd)
            total_turno += qtd
            row += 1
        ws.cell(row=row, column=1, value="Total")
        ws.cell(row=row, column=2, value=total_turno)
        row += 2

    ws.cell(row=row, column=1, value="TOTAL GERAL")
    ws.cell(row=row, column=2, value=total_geral)

    file_id = str(uuid.uuid4())
    file_path = os.path.join(TEMP_FOLDER, f"{file_id}.xlsx")
    wb.save(file_path)

    return jsonify({
        "total": total_geral,
        "turnos": resultados,
        "file_id": file_id
    })

@app.route("/download/<file_id>")
def download(file_id):
    file_path = os.path.join(TEMP_FOLDER, f"{file_id}.xlsx")
    if not os.path.exists(file_path):
        return "Arquivo expirado.", 404
    return send_file(file_path, as_attachment=True, download_name="acordos.xlsx")
